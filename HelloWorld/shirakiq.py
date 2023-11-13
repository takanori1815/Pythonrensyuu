import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter.filedialog

print("作成者:白木 嵩")
wb = Workbook()
ws = wb.active
ws2 = wb.create_sheet()


file_path = tkinter.filedialog.askopenfilename(
    title = "select dat files",             
    filetypes = [("dat files", "*.dat")]    
)

if not file_path:
    print("ファイルが選択されませんでした。")
else:
    # datファイルを読み込む
    df = pd.read_table(file_path, delim_whitespace =True,header=1)
""" print(df) """
for column in df.columns:
    df[column] = df[column].apply(lambda x: float(x) if isinstance(x, str) and 'E' in x else x)

""" print(df) """
for row in dataframe_to_rows(df, index=None, header=True):
    ws.append(row)

for row in dataframe_to_rows(df, index=None, header=True):
    ws2.append(row)

#↓横軸の名前a
#ws["A1"] = "Frequency[GHz]"a

#↓行の数を収得
colum_kazu = ws.max_column 

#↓グラフ作成

#折れ線グラフ
line = LineChart()
line.style = 2

#データの選択

#Y軸の範囲
values = Reference(ws, min_col=2, min_row=1, max_col=colum_kazu , max_row=ws.max_row)
#x軸の範囲
#↑の意味　colは行(横方向)、rowは列(縦方向)を意味する。　つまり1行目で1~102列目を選択している
categories = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)

#↓グラフにデータを追加
line.add_data(values, titles_from_data=True)

#横軸のラベルを設定
line.set_categories(categories)



#↓ラベルの設定
""" x_name = input("x軸の名前は何にします?")
y_name = input("y軸の名前は何にしますか?") """
line.x_axis.title = "Frequency[GHz]"
line.y_axis.title = "Sparameter[dB]"
#↓シートにグラフを追加

ws.add_chart(line, 'A12')

#Excelの名前を変更する。

Excel_name = input("保存する名前は何にしますか？？")
wb.save(Excel_name + ".xlsx")
